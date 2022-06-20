import re
import os
import time
import zipfile
import openpyxl
import config_holder


class DirectoryManager:

    def __init__(self, c: config_holder):
        self.holder = c
        self.config = dict()
        self.get_dict()

    def update_dict(self):
        """
        Updates config field
        :return: None
        """
        self.config = self.holder.read_file()

    def clean_start_folder(self) -> None:
        """
        Sends files from start folder to appropriate folders.
        :return: None
        """

        self.update_dict()
        # przenosze sie do folderu startowego
        os.chdir(self.config['start'][0])

        # w liscie zapisuje pliki ktore sa w folderze start
        start_folder_files = list(os.listdir())

        for file_name in start_folder_files:
            regex = re.split("[.]", file_name)
            name = regex[0]
            if len(regex) == 2:
                extension = regex[1]
            else:
                extension = ''
            for folder_name in self.config['by_names']:
                if folder_name.lower() in name.lower():
                    os.replace(self.config['start'][0] + '/' + file_name, self.config[folder_name][0] + '/' + file_name)
                    self.control_file_size(self.config[folder_name][0], file_name)
                    self.control_number_of_files(self.config[folder_name][0])
                    break

            # jezeli plik nie ma w nazwie nazwy folderu
            if os.path.isfile(self.config['start'][0] + '/' + file_name):
                if extension in self.config['extensions']:
                    os.replace(self.config['start'][0] + '/' + file_name, self.config[extension][0] + '/' + file_name)
                    self.control_file_size(self.config[extension][0], file_name)
                    self.control_number_of_files(self.config[extension][0])
                else:
                    os.replace(self.config['start'][0] + '/' + file_name, self.config['others'][0] + '/' + file_name)
                    self.control_file_size(self.config['others'][0], file_name)
                    self.control_number_of_files(self.config['others'][0])

    def control_number_of_files(self, path: str) -> None:
        """
        Checks the number of files in directory. Removes files if there are too much.
        :param path: Directory path.
        :return: None
        """
        os.chdir(path)
        folder_files = list(os.listdir())
        if len(folder_files) > self.config['max_number_of_files'][0]:
            min_date = time.ctime(os.path.getctime(folder_files[0]))
            file_to_remove = folder_files[0]
            for file_name in folder_files:
                # print(time.ctime(os.path.getctime(file_name)))
                if min_date > time.ctime(os.path.getctime(file_name)):
                    # print('najstarszy = time.ctime(os.path.getctime(file_name))')
                    min_date = time.ctime(os.path.getctime(file_name))
                    file_to_remove = file_name
            # print('usuwam ', file_to_remove)
            os.remove(file_to_remove)

    def control_file_size(self, path: str, file_name: str) -> None:
        """
        Compress the file if it is too big.
        :param path: Directory path.
        :param file_name: Name of the checked file.
        :return: None
        """
        os.chdir(path)
        if os.path.getsize(file_name) > self.config['max_size_of_file'][0]:
            self.file_compress(path, [file_name], file_name.split('.')[0] + '.zip')
            os.remove(file_name)

    def print_folder_names(self) -> None:
        """
        Display folders' names.
        :return: None
        """
        self.update_dict()
        for name in self.config['folder_names']:
            print(name)

    def print_folder_content(self, path: str) -> bool:
        """
        Display contents of directory.
        :param path: Directory path.
        :return: None
        """
        os.chdir(path)
        folder_content = list(os.listdir())
        if len(folder_content) == 0 :
            print('Ten folder jest pusty')
            return False
        for file_name in folder_content:
            print(file_name)

        return True

    def print_file_content(self, path: str, file_name: str) -> None:
        """
        Display .txt or .csv file contents.
        :param path: Directory path.
        :param file_name: Name of displayed file.
        :return: None
        """
        file_name = file_name + '.csv'
        if file_name.split('.')[1] == 'xlsx':
            self.print_xlsx_file(path, file_name)
            return
        os.chdir(path)
        try:
            with open(file_name, "r") as f:
                for line in f:
                    print(line)
        except FileNotFoundError:
            print('Nie ma takiego pliku w tym folderze')
        except UnicodeDecodeError:
            print('Nie udało się odczytać zawartości tego pliku')

    def print_xlsx_file(self, path: str, file_name: str) -> None:
        """
        Display .xlsx file contents.
        :param path: Directory path.
        :param file_name: Name of displayed file.
        :return: None
        """
        xlsx_file = path + '/' + file_name
        wb_obj = openpyxl.load_workbook(xlsx_file)

        sheet = wb_obj.active

        for row in sheet.iter_rows(max_row=sheet.max_row):
            for cell in row:
                if cell.value is None:
                    print(end="\t")
                else:
                    print(cell.value, end="\t")
            print()

    def file_compress(self, path: str, inp_file_names: list[str], out_zip_file: str) -> None:
        """
        Compress selected files.
        :param path: Directory path.
        :param inp_file_names: List of files to compress.
        :param out_zip_file: Name of the zip file.
        :return: None
        """
        os.chdir(path)
        compression = zipfile.ZIP_DEFLATED

        zf = zipfile.ZipFile(out_zip_file, mode="w")

        try:
            for file_to_write in inp_file_names:
                zf.write(file_to_write, file_to_write, compress_type=compression)

        except FileNotFoundError:
            print("Nie znaleziono pliku.")
        finally:
            zf.close()

    def create_folder(self, folder_name: str) -> None:
        """
        Creates folder.
        :param folder_name: New folder name.
        :return: None
        """
        path = os.path.join(self.config['parent_directory'][0], folder_name)
        try:
            os.mkdir(path)
        except FileExistsError:
            print('Folder o tej nazwie istnieje')

    def create_file(self, file_name, text):
        os.chdir(self.config['txt'][0])
        try:
            file = open(file_name, 'x')

            for line in text:
                file.write(line)
        except FileExistsError:
            print("Ten plik już istnieje")
        finally:
            file.close()
            self.control_number_of_files(self.config['txt'][0])
