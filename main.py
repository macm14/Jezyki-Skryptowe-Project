import os
import shutil
import re
import time
import zipfile
import io
import yaml
import openpyxl
import pandas as pd
import matplotlib.pyplot as plt
import numpy as np


def clean_start_folder():
    # przenosze sie do folderu startowego
    os.chdir(config['start'])

    # w liscie zapisuje pliki ktore sa w folderze start
    start_folder_files = list(os.listdir())

    for file_name in start_folder_files:
        regex = re.split("[.]", file_name)
        name = regex[0]
        extension = regex[1]
        for folder_name in config['important_folders']:
            if folder_name.lower() in name.lower():
                os.replace(config['start'] + '/' + file_name, config[folder_name] + '/' + file_name)
                break

        # jezeli plik nie ma w nazwie nazwy folderu
        if os.path.isfile(config['start'] + '/' + file_name):
            if extension in config['extensions']:
                os.replace(config['start'] + '/' + file_name, config[extension] + '/' + file_name)
                control_number_of_files(config[extension])
                control_file_size(config[extension], file_name)
            else:
                os.replace(config['start'] + '/' + file_name, config['others'] + '/' + file_name)
                control_number_of_files(config['others'])


def control_number_of_files(path):
    os.chdir(path)
    folder_files = list(os.listdir())
    if len(folder_files) > config['max_number_of_files']:
        min_date = time.ctime(os.path.getctime(folder_files[0]))
        file_to_remove = folder_files[0]
        for file_name in folder_files:
            if min_date > time.ctime(os.path.getctime(file_name)):
                min_date = time.ctime(os.path.getctime(file_name))
                file_to_remove = file_name
        os.remove(file_to_remove)


def control_file_size(path, file_name):
    os.chdir(path)
    if os.path.getsize(file_name) > config['max_size_of_file']:
        file_compress(path, [file_name], file_name.split('.')[0] + '.zip')
        os.remove(file_name)


def print_folder_names():
    for name in config['folder_names']:
        print(name)


def print_folder_content(path):
    os.chdir(path)
    folder_content = list(os.listdir())

    for file_name in folder_content:
        print(file_name)


def print_file_content(path, file_name):
    if file_name.split('.')[1] == 'xlsx':
        print_xlsx_file(path, file_name)
        return
    os.chdir(path)
    with open(file_name, "r") as f:
        for line in f:
            print(line)


def print_xlsx_file(path, file_name):
    xlsx_file = path + '/' + file_name
    wb_obj = openpyxl.load_workbook(xlsx_file)

    sheet = wb_obj.active

    for row in sheet.iter_rows(max_row=sheet.max_row):
        for cell in row:
            if cell.value is None:
                print(end="\t")
            else:
                print(cell.value, end="\t")
        print()


def get_file_size(file_name):
    return round(os.path.getsize(file_name) / 1000, 2)


def get_folder_size(path):
    size = 0
    os.chdir(path)
    for file_name in list(os.listdir()):
        size += get_file_size(file_name)

    return size


def file_compress(path, inp_file_names, out_zip_file):
    os.chdir(path)
    # Select the compression mode ZIP_DEFLATED for compression
    # or zipfile.ZIP_STORED to just store the file
    compression = zipfile.ZIP_DEFLATED

    zf = zipfile.ZipFile(out_zip_file, mode="w")

    try:
        for file_to_write in inp_file_names:
            # Add file to the zip file
            # first parameter file to zip, second filename in zip
            # print(f' *** Processing file {file_to_write}')
            zf.write(file_to_write, file_to_write, compress_type=compression)

    except FileNotFoundError as e:
        print(f' *** Exception occurred during zip process - {e}')
    finally:
        zf.close()


def all_folders_report():
    folders_list = []
    folders_size = []

    for name in config['folder_names']:
        folders_list.append(name)
        folders_size.append(get_folder_size(config[name]))

    create_plot('Folder name', 'Folder size [KB]', folders_list, folders_size)


def files_report(path):
    os.chdir(path)
    files_list = list(os.listdir())
    files_size = []
    for file_name in files_list:
        files_size.append(get_file_size(file_name))

    create_plot('File name', 'File size [KB]', files_list, files_size)


def create_plot(x_name, y_name, names_list, sizes_list):
    d = {'Name': names_list,
         'Size': sizes_list}
    df = pd.DataFrame(data=d)

    print(df)
    fig, ax = plt.subplots()
    x = np.arange(len(names_list))
    x_bar = ax.bar(x, sizes_list, label='Size')
    ax.set_xticks(x, names_list, rotation=45, fontsize=8)
    ax.set_xlabel(x_name)
    ax.set_ylabel(y_name)
    ax.legend()
    ax.bar_label(x_bar)
    plt.show()


if __name__ == '__main__':

    with open('config.yaml', 'r') as fp:
        config = yaml.safe_load(fp)

    # print(config)
    clean_start_folder()
    user_input = input('Program do zarządzania folderami.\n'
                       'Zrobiłem już porządek w folderach. Co chcesz zrobić następne?\n'
                       '1. Spis folderów.\n'
                       '2. Stwórz własny plik tekstowy.\n'
                       '3. Działania na folderach.\n'
                       '4. Raport dla wszyskitch folderow.\n'
                       '5. Raport dla wybranego folderu\n')
    match user_input:
        case '1':
            print_folder_names()

        case '2':
            user_input = input('Podaj nazwe pliku: ')
            try:
                file = open(user_input, 'x')
                user_input = input('Co chcesz mieć w pliku: ')
                while user_input != '':
                    file.write(user_input)
                    user_input = input()
                    file.write('\n')
                file.close()
            except FileExistsError:
                print("Ten plik już istnieje")

        case '3':
            user_input = input("Na ktorym folderze chcesz operowac:\n 1. text\n 2. data\n")
            chosen_path = ''
            if user_input == '1':
                chosen_path = config['text']
            elif user_input == '2':
                chosen_path = config['data']

            if chosen_path != '':
                print_folder_content(chosen_path)
                user_input = input("Co chcesz zrobic:\n "
                                   "1. Wyswietl zawartosc pliku\n 2. Kompresja plikow\n 3. Raport dla tego folderu.\n")
                match user_input:
                    case '1':
                        user_input = input("Który plik chcesz wyswietlić?\n")
                        print_file_content(chosen_path, user_input)

                    case '2':
                        user_input = input("Które pliki chcesz przenieść do zip?\n")
                        files_to_zip = user_input.split(' ')
                        user_input = input("Podaj nazwę pliku zip\n")
                        file_compress(chosen_path, files_to_zip, user_input + '.zip')

        case '4':
            all_folders_report()

        case '5':
            # user_input = input(  + list(config['folder_names']))
            # print("Dla którego folderu chcesz zrobić raport: ", config['folder_names'])
            print("Dla którego folderu chcesz zrobić raport: ")
            for name in config['folder_names']:
                print(name)

            user_input = input()
            files_report(config[user_input])
